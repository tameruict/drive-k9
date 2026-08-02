"""
Drive Reconcile — Đối chiếu Drive nguồn vs Drive đích
=======================================================
Scan toàn bộ 2 Drive, so sánh và xuất báo cáo:

  MISSING       — có ở nguồn, KHÔNG có ở đích  ← cần copy lại
  EXTRA         — có ở đích,  KHÔNG có ở nguồn ← file lạ / đã xoá ở nguồn
  SIZE_MISMATCH — cùng tên, khác kích thước    ← có thể copy bị lỗi
  FOLDER_ONLY   — folder tồn tại ở nguồn nhưng rỗng ở đích
  OK            — khớp hoàn toàn

Chạy:
  python drive_reconcile.py \
      --source-folder-id FOLDER_ID_NGUON \
      --dest-folder-id   FOLDER_ID_DICH  \
      --token-file token.json \
      [--output report.xlsx] \
      [--max-depth 15] \
      [--ignore-size]   # chỉ kiểm tra tên, bỏ qua kích thước
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

from drive_common import FOLDER_MIME_TYPE
from drive_common import (
    is_shortcut,
    skipped_shortcut_reason,
    without_drive_shortcuts,
)
from drive_common import normalize_drive_name as _normalize_drive_name
from drive_common import normalize_drive_path as _normalize_drive_path

# ─── Constants ────────────────────────────────────────────────────────────────

FOLDER_MIME  = FOLDER_MIME_TYPE
SKIP_MIMES   = set()

STATUS_MISSING        = "MISSING"        # nguồn có, đích không có
STATUS_EXTRA          = "EXTRA"          # đích có, nguồn không có
STATUS_SIZE_MISMATCH  = "SIZE_MISMATCH"  # cùng tên, khác size
STATUS_FOLDER_MISSING = "FOLDER_MISSING" # folder thiếu ở đích
STATUS_OK             = "OK"

# ─── Data classes ─────────────────────────────────────────────────────────────

@dataclass
class DriveFile:
    id:        str
    name:      str
    mime_type: str
    size:      int        # bytes, 0 cho Google Workspace
    path:      str        # đường dẫn đầy đủ trong Drive

    @property
    def is_folder(self) -> bool:
        return self.mime_type == FOLDER_MIME

    @property
    def key(self) -> str:
        """Key chuẩn hoá để so sánh: đường dẫn + tên file."""
        return normalize_path(self.path)


@dataclass
class DiffEntry:
    status:    str
    path:      str
    src_id:    str  = ""
    dst_id:    str  = ""
    src_size:  int  = 0
    dst_size:  int  = 0
    src_mime:  str  = ""
    dst_mime:  str  = ""
    note:      str  = ""


# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    return _normalize_drive_name(name)

def normalize_path(path: str) -> str:
    return _normalize_drive_path(path)

def fmt_size(size: int) -> str:
    if size == 0:
        return "—"
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} TB"

# ─── Drive scanning ───────────────────────────────────────────────────────────

def build_service(token_file: str):
    with open(token_file) as f:
        info = json.load(f)
    creds = Credentials.from_authorized_user_info(info)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


# Thread-local storage so each worker thread gets its own service instance.
# Sharing one service across threads causes SSL corruption (WRONG_VERSION_NUMBER,
# DECRYPTION_FAILED) because httplib2 connections are not thread-safe.
_thread_local = __import__("threading").local()

def _get_thread_service(token_file: str):
    """Return a per-thread Drive service, creating one on first use."""
    svc = getattr(_thread_local, "service", None)
    if svc is None:
        svc = build_service(token_file)
        _thread_local.service = svc
    return svc


def list_children(service, folder_id: str, retries: int = 3) -> list[dict]:
    import time
    items, page_token = [], None
    while True:
        for attempt in range(1, retries + 1):
            try:
                resp = service.files().list(
                    q=f"'{folder_id}' in parents and trashed = false",
                    pageSize=1000,
                    fields="files(id,name,mimeType,size,shortcutDetails),nextPageToken",
                    pageToken=page_token,
                    corpora="allDrives",
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                ).execute()
                items.extend(without_drive_shortcuts(resp.get("files", [])))
                page_token = resp.get("nextPageToken")
                break  # success
            except HttpError as e:
                print(f"  [WARN] list_children({folder_id}) attempt {attempt}: {e}")
                if attempt == retries:
                    return items
                time.sleep(2 ** attempt)
            except Exception as e:
                # Catch SSL / timeout errors
                print(f"  [WARN] list_children({folder_id}) attempt {attempt}: {e}")
                if attempt == retries:
                    return items
                time.sleep(2 ** attempt)
        if not page_token:
            break
    return items


def resolve_shortcut_item(service, item: dict) -> tuple[dict | None, str | None]:
    if not is_shortcut(item):
        return item, None
    return None, skipped_shortcut_reason(item)


def scan_drive(
    service,
    folder_id: str,
    root_path: str = "",
    max_depth: int = 15,
    depth: int = 0,
    results: Optional[dict[str, DriveFile]] = None,
    folders: Optional[dict[str, DriveFile]] = None,
) -> tuple[dict[str, DriveFile], dict[str, DriveFile]]:
    """
    Đệ quy scan toàn bộ folder.
    Trả về (files_index, folders_index) với key là normalized path.
    """
    if results is None:
        results = {}
    if folders is None:
        folders = {}
    if depth > max_depth:
        print(f"  [WARN] Vượt max_depth tại: {root_path}")
        return results, folders

    children = list_children(service, folder_id)

    for item in children:
        if is_shortcut(item):
            continue
        resolved, error = resolve_shortcut_item(service, item)
        if error:
            print(f"  [WARN] {error}")
            continue
        item = resolved or item
        mime  = item.get("mimeType", "")
        name  = item.get("name", "")
        iid   = item.get("_target_id", item.get("id", ""))
        size  = int(item.get("size", 0))
        path  = f"{root_path}/{name}" if root_path else name

        if mime in SKIP_MIMES:
            continue

        df = DriveFile(id=iid, name=name, mime_type=mime, size=size, path=path)

        if mime == FOLDER_MIME:
            folders[df.key] = df
            scan_drive(
                service,
                item.get("_list_folder_id", iid),
                path,
                max_depth,
                depth + 1,
                results,
                folders,
            )
        else:
            results[df.key] = df

    return results, folders


def scan_parallel(
    service,
    folder_id: str,
    max_depth: int = 15,
    workers: int = 4,
    token_file: str = "token.json",
) -> tuple[dict[str, DriveFile], dict[str, DriveFile]]:
    """
    Scan folder với đa luồng ở tầng con đầu tiên để tăng tốc.
    Mỗi thread dùng service riêng để tránh lỗi SSL (WRONG_VERSION_NUMBER).
    """
    # Lấy folder con tầng 1
    top = list_children(service, folder_id)
    resolved_top = []
    for item in top:
        if is_shortcut(item):
            continue
        resolved, error = resolve_shortcut_item(service, item)
        if error:
            print(f"  [WARN] {error}")
            continue
        resolved_top.append(resolved or item)
    top_folders = [i for i in resolved_top if i.get("mimeType") == FOLDER_MIME]
    top_files   = [i for i in resolved_top if i.get("mimeType") != FOLDER_MIME
                   and i.get("mimeType") not in SKIP_MIMES]

    all_files:   dict[str, DriveFile] = {}
    all_folders: dict[str, DriveFile] = {}

    # File rời ở root
    for item in top_files:
        df = DriveFile(
            id=item.get("_target_id", item["id"]), name=item["name"],
            mime_type=item.get("mimeType", ""),
            size=int(item.get("size", 0)),
            path=item["name"],
        )
        all_files[df.key] = df

    def _scan_subfolder(folder_item: dict):
        # Each worker thread gets its own Drive service — not thread-safe to share.
        thread_svc = _get_thread_service(token_file)
        name = folder_item["name"]
        fid  = folder_item["id"]
        scan_id = folder_item.get("_list_folder_id", fid)
        df   = DriveFile(id=fid, name=name, mime_type=FOLDER_MIME, size=0, path=name)
        sub_files, sub_folders = scan_drive(
            thread_svc, scan_id, root_path=name, max_depth=max_depth
        )
        return df, sub_files, sub_folders

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_scan_subfolder, f): f for f in top_folders}
        for future in as_completed(futures):
            try:
                df, sub_files, sub_folders = future.result()
                all_folders[df.key] = df
                all_files.update(sub_files)
                all_folders.update(sub_folders)
            except Exception as e:
                print(f"  [ERROR] scan subfolder: {e}")

    return all_files, all_folders


# ─── Diff logic ───────────────────────────────────────────────────────────────

def diff_drives(
    src_files:   dict[str, DriveFile],
    dst_files:   dict[str, DriveFile],
    src_folders: dict[str, DriveFile],
    dst_folders: dict[str, DriveFile],
    ignore_size: bool = False,
) -> list[DiffEntry]:
    diffs: list[DiffEntry] = []

    # 1. File có ở nguồn
    for key, src in src_files.items():
        dst = dst_files.get(key)

        if dst is None:
            diffs.append(DiffEntry(
                status=STATUS_MISSING,
                path=src.path,
                src_id=src.id, src_size=src.size, src_mime=src.mime_type,
                note="Không tìm thấy ở Drive đích",
            ))
        elif not ignore_size and src.size > 0 and dst.size > 0 and src.size != dst.size:
            diffs.append(DiffEntry(
                status=STATUS_SIZE_MISMATCH,
                path=src.path,
                src_id=src.id, dst_id=dst.id,
                src_size=src.size, dst_size=dst.size,
                src_mime=src.mime_type, dst_mime=dst.mime_type,
                note=f"Size: nguồn {fmt_size(src.size)} ≠ đích {fmt_size(dst.size)}",
            ))
        else:
            diffs.append(DiffEntry(
                status=STATUS_OK,
                path=src.path,
                src_id=src.id, dst_id=dst.id,
                src_size=src.size, dst_size=dst.size,
            ))

    # 2. File chỉ có ở đích (file lạ)
    for key, dst in dst_files.items():
        if key not in src_files:
            diffs.append(DiffEntry(
                status=STATUS_EXTRA,
                path=dst.path,
                dst_id=dst.id, dst_size=dst.size, dst_mime=dst.mime_type,
                note="Chỉ tồn tại ở Drive đích (nguồn không có)",
            ))

    # 3. Folder thiếu ở đích
    for key, src_f in src_folders.items():
        if key not in dst_folders:
            diffs.append(DiffEntry(
                status=STATUS_FOLDER_MISSING,
                path=src_f.path,
                src_id=src_f.id,
                note="Folder có ở nguồn nhưng không có ở đích",
            ))

    return diffs


# ─── Report ───────────────────────────────────────────────────────────────────

def print_summary(diffs: list[DiffEntry], elapsed: float):
    counts = defaultdict(int)
    for d in diffs:
        counts[d.status] += 1

    total = len(diffs)
    ok    = counts[STATUS_OK]

    print("\n" + "=" * 60)
    print("📋 KẾT QUẢ ĐỐI CHIẾU")
    print("=" * 60)
    print(f"  ✅ Khớp hoàn toàn          : {ok:>6} / {total}")
    print(f"  🚫 MISSING  (thiếu ở đích) : {counts[STATUS_MISSING]:>6}")
    print(f"  ⚠️  SIZE_MISMATCH           : {counts[STATUS_SIZE_MISMATCH]:>6}")
    print(f"  📁 FOLDER_MISSING          : {counts[STATUS_FOLDER_MISSING]:>6}")
    print(f"  ➕ EXTRA (lạ ở đích)       : {counts[STATUS_EXTRA]:>6}")
    print(f"  ⏱️  Thời gian scan          : {elapsed:.1f}s")
    print("=" * 60)

    issues = [d for d in diffs if d.status != STATUS_OK]
    if not issues:
        print("\n🎉 Không có sai lệch nào — 2 Drive đồng bộ hoàn toàn!")
        return

    print(f"\n⚠️  {len(issues)} sai lệch cần xử lý:\n")

    # Nhóm theo status
    for status in [STATUS_MISSING, STATUS_SIZE_MISMATCH, STATUS_FOLDER_MISSING, STATUS_EXTRA]:
        group = [d for d in issues if d.status == status]
        if not group:
            continue
        labels = {
            STATUS_MISSING:        "🚫 MISSING — có ở nguồn, thiếu ở đích",
            STATUS_SIZE_MISMATCH:  "⚠️  SIZE_MISMATCH — cùng tên, khác kích thước",
            STATUS_FOLDER_MISSING: "📁 FOLDER_MISSING — folder thiếu ở đích",
            STATUS_EXTRA:          "➕ EXTRA — chỉ có ở đích",
        }
        print(f"  {labels[status]} ({len(group)} mục):")
        for d in group[:15]:    # in tối đa 15 dòng mỗi loại
            print(f"    • {d.path}")
            if d.note:
                print(f"      → {d.note}")
        if len(group) > 15:
            print(f"    ... và {len(group) - 15} mục khác (xem file báo cáo)")
        print()


def export_json(diffs: list[DiffEntry], path: str):
    data = [
        {
            "status":   d.status,
            "path":     d.path,
            "src_id":   d.src_id,
            "dst_id":   d.dst_id,
            "src_size": d.src_size,
            "dst_size": d.dst_size,
            "src_mime": d.src_mime,
            "dst_mime": d.dst_mime,
            "note":     d.note,
        }
        for d in diffs
    ]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"📄 JSON: {path}")


def export_excel(diffs: list[DiffEntry], path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("[WARN] openpyxl chưa cài. Bỏ qua export Excel. (pip install openpyxl)")
        return

    STATUS_COLORS = {
        STATUS_MISSING:        "FFCCCC",   # đỏ nhạt
        STATUS_SIZE_MISMATCH:  "FFF0CC",   # cam nhạt
        STATUS_FOLDER_MISSING: "FFE5CC",   # cam nhạt hơn
        STATUS_EXTRA:          "E5CCFF",   # tím nhạt
        STATUS_OK:             "CCFFCC",   # xanh nhạt
    }

    wb = openpyxl.Workbook()

    # ── Sheet 1: Toàn bộ ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Toàn bộ"
    headers = ["Trạng thái", "Đường dẫn", "Size nguồn", "Size đích", "Ghi chú",
               "ID nguồn", "ID đích"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    for d in diffs:
        row = [
            d.status, d.path,
            fmt_size(d.src_size), fmt_size(d.dst_size),
            d.note, d.src_id, d.dst_id,
        ]
        ws.append(row)
        color = STATUS_COLORS.get(d.status, "FFFFFF")
        fill  = PatternFill(fill_type="solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40

    # ── Sheet 2: Chỉ sai lệch ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sai lệch")
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    issues = [d for d in diffs if d.status != STATUS_OK]
    for d in issues:
        row = [
            d.status, d.path,
            fmt_size(d.src_size), fmt_size(d.dst_size),
            d.note, d.src_id, d.dst_id,
        ]
        ws2.append(row)
        color = STATUS_COLORS.get(d.status, "FFFFFF")
        fill  = PatternFill(fill_type="solid", fgColor=color)
        for cell in ws2[ws2.max_row]:
            cell.fill = fill

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["E"].width = 40

    # ── Sheet 3: Chỉ MISSING ─────────────────────────────────────────────────
    missing = [d for d in diffs if d.status == STATUS_MISSING]
    if missing:
        ws3 = wb.create_sheet("MISSING")
        ws3.append(["Đường dẫn", "Size nguồn", "MIME type", "ID nguồn"])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        for d in missing:
            ws3.append([d.path, fmt_size(d.src_size), d.src_mime, d.src_id])
            fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
            for cell in ws3[ws3.max_row]:
                cell.fill = fill
        ws3.column_dimensions["A"].width = 60

    wb.save(path)
    print(f"📊 Excel: {path}  ({len(diffs)} dòng, {len(issues)} sai lệch)")


def export_missing_json(diffs: list[DiffEntry], path: str):
    """Xuất riêng danh sách file MISSING để dễ dùng với script copy lại."""
    missing = [
        {"path": d.path, "src_id": d.src_id, "src_size": d.src_size, "src_mime": d.src_mime}
        for d in diffs if d.status == STATUS_MISSING
    ]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(missing, f, ensure_ascii=False, indent=2)
    print(f"📄 Missing list: {path}  ({len(missing)} file)")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Đối chiếu Drive nguồn vs Drive đích và xuất báo cáo sai lệch."
    )
    p.add_argument("--source-folder-id", required=True,  help="Folder ID Drive nguồn")
    p.add_argument("--dest-folder-id",   required=True,  help="Folder ID Drive đích")
    p.add_argument("--token-file",       default="token.json")
    p.add_argument("--output",           default="drive_reconcile_report.xlsx",
                   help="File báo cáo Excel (mặc định: drive_reconcile_report.xlsx)")
    p.add_argument("--json-output",      default="drive_reconcile_report.json")
    p.add_argument("--missing-output",   default="drive_missing.json",
                   help="File chỉ chứa danh sách MISSING để copy lại")
    p.add_argument("--max-depth",        type=int, default=15)
    p.add_argument("--scan-workers",     type=int, default=4,
                   help="Số thread dùng khi scan (mặc định: 4)")
    p.add_argument("--ignore-size",      action="store_true",
                   help="Chỉ đối chiếu tên file, bỏ qua kích thước")
    p.add_argument("--no-ok",            action="store_true",
                   help="Không ghi dòng OK vào báo cáo (chỉ ghi sai lệch)")
    return p.parse_args()


def main():
    args = parse_args()

    print("=" * 60)
    print("  DRIVE RECONCILE — Đối chiếu 2 Drive")
    print(f"  Nguồn : {args.source_folder_id}")
    print(f"  Đích   : {args.dest_folder_id}")
    print(f"  Thời điểm: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 60)

    service = build_service(args.token_file)

    # ── Scan nguồn ────────────────────────────────────────────────────────────
    print("\n📡 Đang scan Drive NGUỒN...")
    t0 = __import__("time").monotonic()
    src_files, src_folders = scan_parallel(
        service, args.source_folder_id,
        max_depth=args.max_depth, workers=args.scan_workers,
        token_file=args.token_file,
    )
    t_src = __import__("time").monotonic() - t0
    print(f"   → {len(src_files):,} file, {len(src_folders):,} folder  ({t_src:.1f}s)")

    # ── Scan đích ─────────────────────────────────────────────────────────────
    print("\n📡 Đang scan Drive ĐÍCH...")
    t1 = __import__("time").monotonic()
    dst_files, dst_folders = scan_parallel(
        service, args.dest_folder_id,
        max_depth=args.max_depth, workers=args.scan_workers,
        token_file=args.token_file,
    )
    t_dst = __import__("time").monotonic() - t1
    print(f"   → {len(dst_files):,} file, {len(dst_folders):,} folder  ({t_dst:.1f}s)")

    # ── Đối chiếu ─────────────────────────────────────────────────────────────
    print("\n🔍 Đang đối chiếu...")
    t2 = __import__("time").monotonic()
    diffs = diff_drives(
        src_files, dst_files, src_folders, dst_folders,
        ignore_size=args.ignore_size,
    )

    if args.no_ok:
        diffs = [d for d in diffs if d.status != STATUS_OK]

    elapsed = __import__("time").monotonic() - t0

    # ── In kết quả ────────────────────────────────────────────────────────────
    print_summary(diffs, elapsed)

    # ── Xuất báo cáo ──────────────────────────────────────────────────────────
    print("\n💾 Đang xuất báo cáo...")
    export_json(diffs, args.json_output)
    export_excel(diffs, args.output)
    export_missing_json(diffs, args.missing_output)

    # Exit code: 0 = hoàn toàn khớp, 1 = có sai lệch
    has_issues = any(d.status != STATUS_OK for d in diffs)
    sys.exit(1 if has_issues else 0)


if __name__ == "__main__":
    main()
